"""A run's answers, one table per recipe, in the shape the original Junior used.

Taken from the pre-rewrite pipeline's own run outputs — the ones used for the
100-patient experiment:

    Run_Outputs/exp5/20260311_120527/gpt5/date_of_birth_results_20260311_120527.xlsx

One file per variable, every file stamped with the run, all in one folder, each sheet
one row per patient with `var_name, ok, <every field the recipe produced>, patient_id`.

The correction worth keeping: an earlier version of this dropped the certainty scores,
the rationale and the evidence chunk ids as "the model's commentary". They are not.
A value beside its verbatim quote beside the chunk it came from is what makes a row
checkable, which is the entire claim this pipeline makes.
"""
from __future__ import annotations

import csv
import io
import json
from pathlib import Path

from jr_pipeline.runtime_infrastructure.data_directory_layout_and_safe_writes import (
    recipe_table_name,
    run_output_dir,
)
from jr_pipeline.runtime_infrastructure.values_table import (
    read_run_values,
    recipe_tables,
    write_values_tables,
)

RUN = "20260820_105033_b74d"


def _a_result(run_root: Path, patient: str, variable: str, data: dict, ok: bool = True):
    d = run_root / "patients" / patient / "extract" / variable
    d.mkdir(parents=True, exist_ok=True)
    (d / "result.json").write_text(
        json.dumps({"payload": {"ok": ok, "variable": variable, "data": data}}),
        encoding="utf-8",
    )


def _rows(text: str) -> list[dict]:
    return list(csv.DictReader(io.StringIO(text)))


def _sheet(path: Path) -> list[dict]:
    """A written workbook back as rows, so a test reads what a clinician opens."""
    import openpyxl

    ws = openpyxl.load_workbook(path).active
    header = [c.value for c in ws[1]]
    return [dict(zip(header, row, strict=True)) for row in ws.iter_rows(min_row=2, values_only=True)]


def test_the_evidence_travels_with_the_value(tmp_path):
    """The whole point of this pipeline is that a value can be checked. A table of bare
    values is the one thing a clinical extraction must not hand back."""
    _a_result(tmp_path, "P1", "date_of_birth", {
        "date_of_birth": "1973-01-02", "dob_certainty": 100, "dob_evidence": "explicit",
        "dob_evidence_chunk_id": "P1:demographics:0:TABLE", "rationale": "stated in the row",
    })

    row = _rows(recipe_tables(tmp_path)["date_of_birth"])[0]

    assert row["date_of_birth"] == "1973-01-02"
    for kept in ("dob_certainty", "dob_evidence", "dob_evidence_chunk_id", "rationale"):
        assert row[kept], f"{kept} was dropped"


def test_one_table_per_recipe_not_one_wide_table(tmp_path):
    """Each recipe has its own fields. Joining them into a single table produced sparse
    columns nobody asked for, and 285 of them once a row-mapped recipe was included."""
    _a_result(tmp_path, "P1", "date_of_birth", {"date_of_birth": "1973-01-02"})
    _a_result(tmp_path, "P1", "stage", {"stage_at_diagnosis": {"overall_group": "IIA"}})

    tables = recipe_tables(tmp_path)

    assert set(tables) == {"date_of_birth", "stage"}
    assert "stage_at_diagnosis.overall_group" in _rows(tables["stage"])[0]


def test_every_row_says_whose_it_is_and_whether_to_trust_it(tmp_path):
    _a_result(tmp_path, "P1", "date_of_birth", {"date_of_birth": "1973-01-02"})

    header = list(_rows(recipe_tables(tmp_path)["date_of_birth"])[0])

    assert header[:3] == ["patient_id", "var_name", "ok"]


def test_a_failed_variable_is_in_the_table_marked_failed(tmp_path):
    """Dropping it would make a broken extraction indistinguishable from a chart with
    nothing in it — the distinction the whole pipeline exists to preserve."""
    _a_result(tmp_path, "P1", "stage", {"clinical_tnm": {"cT": None}}, ok=False)

    row = _rows(recipe_tables(tmp_path)["stage"])[0]

    # TRUE/FALSE, not Python's True/False: read.csv gives a character column for the
    # latter and `subset(d, ok)` errors outright.
    assert row["ok"] == "FALSE"


def test_a_row_mapped_recipe_gets_one_row_per_record(tmp_path):
    """pathology_table makes one model call per pathology report and answers with a
    list, so its natural table is the table it came from."""
    _a_result(tmp_path, "P1", "pathology_table", {
        "n_source_rows": 2,
        "rows": [{"specimen": "left breast", "cancer_status": "yes"},
                 {"specimen": "blood", "cancer_status": "no"}],
    })

    rows = _rows(recipe_tables(tmp_path)["pathology_table"])

    assert len(rows) == 2
    assert [r["row"] for r in rows] == ["0", "1"]
    assert rows[0]["specimen"] == "left breast"
    # The patient-level fields ride along, so a row stands on its own.
    assert rows[0]["n_source_rows"] == "2"


def test_a_row_mapped_recipe_that_found_nothing_still_appears(tmp_path):
    _a_result(tmp_path, "P1", "pathology_table", {"n_source_rows": 0, "rows": []})

    assert len(_rows(recipe_tables(tmp_path)["pathology_table"])) == 1


def test_the_files_are_named_and_placed_like_the_originals(tmp_path):
    """<variable>_results_<timestamp>.csv, in one folder, because these leave: they get
    downloaded, emailed, and opened next to somebody else's."""
    run = tmp_path / RUN
    _a_result(run, "P1", "date_of_birth", {"date_of_birth": "1973-01-02"})

    written = write_values_tables(run)

    assert [p.name for p in written] == ["date_of_birth_results.xlsx", "run_metadata.xlsx"]
    assert written[0].parent == run_output_dir(run)
    assert recipe_table_name(RUN, "stage") == "stage_results.xlsx"


def test_two_runs_answers_do_not_land_on_each_other(tmp_path):
    """Asking the same recipes of one embedded cohort twice is the ordinary way to work
    here, and the run id used to be trimmed to its first fifteen characters — the
    timestamp, not the hex suffix that tells two runs in the same second apart.
    compare_llm_models starts several on purpose, named <run id>__compare__<model>, so
    comparing three endpoints put four sets of answers on one path and kept the last."""
    receipts = tmp_path / "CONTAINS_PHI" / "pipeline_run_receipts"
    run_ids = [RUN] + [f"{RUN}__compare__{m}" for m in ("gpt5", "claude_opus", "llama3")]
    for run_id in run_ids:
        _a_result(receipts / run_id, "P1", "stage", {"overall_group": "IIA"})
        write_values_tables(receipts / run_id)

    landed = {run_output_dir(receipts / run_id) / recipe_table_name(run_id, "stage")
              for run_id in run_ids}

    assert len(landed) == 4, "two runs wrote their answers to one path"
    assert all(path.is_file() for path in landed)


def test_a_run_finds_its_answers_beside_every_other_runs(tmp_path):
    """One dated folder per run, all siblings. They used to sit inside the run receipt
    in a folder called run_output_all, so every run had a folder of that same name
    buried under its own id, and comparing two runs meant two deep paths whose last two
    segments were identical."""
    receipts = tmp_path / "CONTAINS_PHI" / "pipeline_run_receipts"
    for run_id in (RUN, "20260821_094512_7c1a"):
        _a_result(receipts / run_id, "P1", "date_of_birth", {"date_of_birth": "1973-01-02"})
        write_values_tables(receipts / run_id)

    answers = tmp_path / "CONTAINS_PHI" / "answers"

    assert sorted(p.name for p in answers.iterdir()) == [RUN, "20260821_094512_7c1a"]
    assert (answers / RUN / "date_of_birth_results.xlsx").is_file()
    # The answers are chart values, so they stay on the PHI side of the project.
    assert "CONTAINS_PHI" in str(answers)


def test_a_run_id_that_would_escape_the_folder_is_refused(tmp_path):
    """The run id names a directory AND is stamped into a downloadable filename, so it
    reaches disk twice. validate_patient_id has guarded the other such id since the
    start; this one was joined into both paths unchecked."""
    import pytest

    for bad in ("../../etc", "a/b", "", "."):
        with pytest.raises(ValueError):
            recipe_table_name(bad, "stage")


def test_the_answers_are_not_loose_among_the_machinery(tmp_path):
    """A run folder holds fifteen entries of which the answers are the point; the rest
    is a response cache, an hmac secret, state fragments and two jsonl logs."""
    run = tmp_path / RUN
    _a_result(run, "P1", "date_of_birth", {"date_of_birth": "1973-01-02"})

    write_values_tables(run)

    assert not list(run.glob("*.xlsx")), "an answer table was left at the run root"
    assert list(run_output_dir(run).glob("*.xlsx"))


def test_a_run_with_nothing_extracted_writes_nothing(tmp_path):
    """Rather than an empty folder that looks like output you failed to find."""
    assert write_values_tables(tmp_path) == []
    assert not run_output_dir(tmp_path).exists()


def test_it_reads_what_is_on_disk(tmp_path):
    """A half-finished run reports what it has."""
    _a_result(tmp_path, "P1", "date_of_birth", {"date_of_birth": "1973-01-02"})
    _a_result(tmp_path, "P2", "date_of_birth", {"date_of_birth": "1980-05-05"})

    assert len(read_run_values(tmp_path)["date_of_birth"]) == 2


# --- the companion data this pipeline records and the original did not --------------

def test_a_failed_row_says_why(tmp_path):
    """The original carried var_name and ok and nothing else about trust, so an
    ok=False row was a dead end: you could see it had failed and not why."""
    d = tmp_path / "patients" / "P1" / "extract" / "stage"
    d.mkdir(parents=True)
    (d / "result.json").write_text(json.dumps({
        "payload": {
            "ok": False, "variable": "stage",
            "data": {"clinical_tnm": {"cT": "T1c", "evidence_chunk_id": None}},
            "errors": ["these values carry no source span"],
            "unprovenanced_value_paths": ["data.clinical_tnm"],
            "provenance_rejected": [
                {"path": "data.clinical_tnm.evidence_chunk_id",
                 "chunk_id": "P1:clinical_note:368:0", "step": "finalize"},
            ],
            "recipe": {"name": "stage", "version": "v1"},
        },
        "produced_by": {"recipe_hash": "sha256:r", "provider_config_hash": "sha256:m",
                        "run_id": "20260820_105033_b74d", "created_at": "2026-08-20T18:30:43Z"},
    }), encoding="utf-8")

    row = _rows(recipe_tables(tmp_path)["stage"])[0]

    assert "no source span" in row["why_it_failed"]
    assert row["values_with_no_evidence"] == "data.clinical_tnm"
    assert row["run_id"] == "20260820_105033_b74d"


def test_a_rejected_citation_is_on_the_row_and_says_which_value(tmp_path):
    """The flagship signal of this pipeline, and it was mislabelled. The check compares
    a cited chunk against the chunks shown to THIS step; it cannot tell an invented id
    from a real one the model reached for out of window, and in the reference run it
    fired on a real note the same run cited approvingly elsewhere. The column also
    dropped the path, leaving a reader with a chunk id and a dozen values and no way to
    tell which one was being complained about."""
    d = tmp_path / "patients" / "P1" / "extract" / "stage"
    d.mkdir(parents=True)
    (d / "result.json").write_text(json.dumps({
        "payload": {
            "ok": False, "variable": "stage", "data": {"clinical_tnm": {"cT": "T1c"}},
            "provenance_rejected": [
                {"path": "data.clinical_tnm.evidence_chunk_id",
                 "chunk_id": "P1:clinical_note:368:0", "step": "finalize"},
                {"path": "data.stage_at_diagnosis.evidence_chunk_id",
                 "chunk_id": "P1:clinical_note:368:0", "step": "finalize"},
            ],
        },
        "produced_by": {},
    }), encoding="utf-8")

    row = _rows(recipe_tables(tmp_path)["stage"])[0]

    assert "citations_the_model_invented" not in row
    assert row["cited_a_chunk_not_shown_to_this_step"] == (
        "data.clinical_tnm.evidence_chunk_id=P1:clinical_note:368:0; "
        "data.stage_at_diagnosis.evidence_chunk_id=P1:clinical_note:368:0")


def test_a_value_whose_pointer_was_thrown_out_still_gets_an_evidence_column(tmp_path):
    """The gate nulls the pointer, so the row no longer holds one and the value ended
    up with no evidence columns at all — stage_at_diagnosis.overall_group = I sat in
    the reference run with nothing beside it, which reads as "nobody asked" rather
    than "the citation was thrown out"."""
    d = tmp_path / "patients" / "P1" / "extract" / "stage"
    d.mkdir(parents=True)
    (d / "result.json").write_text(json.dumps({
        "payload": {
            "ok": True, "variable": "stage",
            "data": {"stage_at_diagnosis": {"overall_group": "I", "evidence_chunk_id": None}},
            "provenance_rejected": [
                {"path": "data.stage_at_diagnosis.evidence_chunk_id",
                 "chunk_id": "P1:clinical_note:368:0", "step": "finalize"},
            ],
        },
        "produced_by": {},
    }), encoding="utf-8")

    row = _rows(recipe_tables(tmp_path)["stage"])[0]

    assert row["stage_at_diagnosis.overall_group"] == "I"
    assert row["stage_at_diagnosis.evidence_text"] == (
        "(citation rejected: P1:clinical_note:368:0 was not shown to this step)")


def test_which_model_produced_the_row(tmp_path):
    """The original's llm_key. Identical on every row of a file, so it is stated once
    in run_metadata.csv and joined back on run_id + var_name; without it anywhere a
    pooled table cannot tell two models apart."""
    _a_result(tmp_path, "P1", "date_of_birth", {"date_of_birth": "1973-01-02"})
    d = tmp_path / "patients" / "P1" / "extract" / "date_of_birth" / "result.json"
    envelope = json.loads(d.read_text())
    envelope["produced_by"] = {"provider_config_hash": "sha256:model-a",
                               "run_id": RUN, "recipe_hash": "sha256:r"}
    d.write_text(json.dumps(envelope), encoding="utf-8")
    write_values_tables(tmp_path)

    metadata = _sheet(run_output_dir(tmp_path) / "run_metadata.xlsx")[0]

    assert metadata["model"] == "sha256:model-a"
    assert metadata["var_name"] == "date_of_birth"
    assert metadata["run_id"] == RUN
    # The join key stays on the row; only what describes the machinery leaves it.
    row = _rows(recipe_tables(tmp_path)["date_of_birth"])[0]
    assert row["run_id"] == RUN
    assert "model" not in row and "recipe_hash" not in row


def test_the_answers_still_come_before_the_companion_columns(tmp_path):
    """A reader came for the values. Trust and provenance are what they consult after,
    and what says whose row it is comes first of all."""
    _a_result(tmp_path, "P1", "date_of_birth", {"date_of_birth": "1973-01-02"})

    header = list(_rows(recipe_tables(tmp_path)["date_of_birth"])[0])

    assert header.index("run_id") < header.index("date_of_birth")
    assert header.index("date_of_birth") < header.index("why_it_failed")
    assert header.index("why_it_failed") < header.index("extracted_at")


def test_every_row_of_a_row_mapped_recipe_carries_them(tmp_path):
    """Each record is judged and traced the same way, so each row repeats them."""
    _a_result(tmp_path, "P1", "pathology_table", {
        "rows": [{"specimen": "left breast"}, {"specimen": "blood"}]})

    rows = _rows(recipe_tables(tmp_path)["pathology_table"])

    assert len(rows) == 2
    assert all("why_it_failed" in row and "run_id" in row for row in rows)


# --- the passage, beside the value ---------------------------------------------------

def _a_passage(run_root: Path, patient: str, variable: str, step: str, blocks: list[dict]):
    d = run_root / "patients" / patient / "prepared_evidence_text" / variable / step
    d.mkdir(parents=True, exist_ok=True)
    (d / "evidence_blocks.json").write_text(json.dumps(blocks), encoding="utf-8")


def test_the_cited_passage_is_in_the_row(tmp_path):
    """The pipeline cannot know whether a value is right — there is no ground truth
    here. What it can do is put the passage next to the value so the decision is a
    person's and takes five seconds rather than eight steps and two parquet reads.

    Modelled on the real failure: a death date cited to a note about bladder infections,
    which no column could have judged and any reader spots instantly."""
    _a_result(tmp_path, "P1", "date_of_death", {
        "date_of_death": "2025-03-27", "vital_status": "dead", "confidence": 100,
        "evidence_chunk_id": "P1:clinical_note:77:3",
    })
    _a_passage(tmp_path, "P1", "date_of_death", "death_search", [{
        "chunk_id": "P1:clinical_note:77:3",
        "text": "SEEK MEDICAL CARE IF: You have back pain. You develop a fever.",
        "source_file": "clinical_note.csv", "record_index": 77,
        "document_date": "2017-11-09", "doc_type": "Progress Note, Outpatient",
    }])

    row = _rows(recipe_tables(tmp_path)["date_of_death"])[0]

    assert "back pain" in row["evidence_text"]
    assert "2017-11-09" in row["evidence_source"]
    assert "clinical_note.csv" in row["evidence_source"]
    assert "record 77" in row["evidence_source"]


def test_the_passage_column_is_named_after_the_pointer_it_belongs_to(tmp_path):
    """A row can carry several values, each with its own citation. An unlabelled quote
    would leave a reader guessing which value it supports."""
    _a_result(tmp_path, "P1", "date_of_birth", {
        "date_of_birth": "1973-01-02", "dob_evidence_chunk_id": "P1:demographics:0:TABLE",
    })
    _a_passage(tmp_path, "P1", "date_of_birth", "demographics_grab", [{
        "chunk_id": "P1:demographics:0:TABLE", "text": "date_of_birth: 1973-01-02",
        "source_file": "demographics.csv",
    }])

    row = _rows(recipe_tables(tmp_path)["date_of_birth"])[0]

    assert row["dob_evidence_text"] == "date_of_birth: 1973-01-02"


def test_a_pointer_to_a_passage_never_shown_says_so(tmp_path):
    """Different from "we did not keep the text". This is the case the provenance gate
    already flags, and the row should agree with it rather than going blank."""
    _a_result(tmp_path, "P1", "stage", {
        "stage_at_diagnosis": {"overall_group": "I",
                               "evidence_chunk_id": "P1:clinical_note:368:0"},
    })

    row = _rows(recipe_tables(tmp_path)["stage"])[0]

    assert "not among those shown" in row["stage_at_diagnosis.evidence_text"]


# --- silent data loss ----------------------------------------------------------------

def test_a_list_of_evidence_pointers_is_not_dropped(tmp_path):
    """_flatten skipped every list, so a recipe whose provenance is a list lost the
    whole column — the one thing this file exists to carry."""
    _a_result(tmp_path, "P1", "stage", {
        "overall_group": "IIA",
        "supporting_chunk_ids": ["P1:path:1:0", "P1:note:9:2"],
    })

    row = _rows(recipe_tables(tmp_path)["stage"])[0]

    assert row["supporting_chunk_ids"] == "P1:path:1:0; P1:note:9:2"


def test_two_record_lists_are_not_silently_resolved_by_sort_order(tmp_path):
    """It took whichever field sorted first as the row axis and discarded the other. A
    recipe that grew a second list would have had every row change meaning, with no
    error anywhere."""
    _a_result(tmp_path, "P1", "v", {
        "medications": [{"name": "tamoxifen"}],
        "rows": [{"specimen": "left breast"}, {"specimen": "blood"}],
    })

    rows = _rows(recipe_tables(tmp_path)["v"])

    assert len(rows) == 1
    assert rows[0]["row_axis_ambiguous"] == "medications; rows"
    # Neither list is lost: both are rendered rather than one being chosen and the
    # other dropped.
    assert "tamoxifen" in rows[0]["medications"]
    assert "blood" in rows[0]["rows"]


# --- the file that leaves the folder -------------------------------------------------

def test_the_write_is_atomic(tmp_path):
    """The one file that gets downloaded and emailed was the only writer in this repo
    not using tmp + os.replace."""
    run = tmp_path / RUN
    _a_result(run, "P1", "date_of_birth", {"date_of_birth": "1973-01-02"})

    write_values_tables(run)

    assert not list(run_output_dir(run).glob(".*tmp")), "a temp file was left behind"


def test_a_table_for_a_dropped_variable_is_removed(tmp_path):
    """Left in place it keeps being counted as one of "the answers" by the run index,
    and keeps being read by whoever opens the folder."""
    run = tmp_path / RUN
    _a_result(run, "P1", "date_of_birth", {"date_of_birth": "1973-01-02"})
    _a_result(run, "P1", "retired_variable", {"value": "x"})
    write_values_tables(run)
    assert len(list(run_output_dir(run).glob("*.xlsx"))) == 3  # two recipes + metadata

    import shutil
    shutil.rmtree(run / "patients" / "P1" / "extract" / "retired_variable")
    write_values_tables(run)

    remaining = sorted(p.name for p in run_output_dir(run).glob("*.xlsx"))
    assert remaining == ["date_of_birth_results.xlsx", "run_metadata.xlsx"]


def test_booleans_are_readable_by_r(tmp_path):
    """Python's True/False makes read.csv give a character column, so sum(d$ok) and
    subset(d, ok) both error. The pipeline this replaced stored a real boolean."""
    _a_result(tmp_path, "P1", "date_of_birth", {"date_of_birth": "1973-01-02"})

    assert _rows(recipe_tables(tmp_path)["date_of_birth"])[0]["ok"] == "TRUE"


def test_a_row_says_whether_anything_was_actually_found(tmp_path):
    """`ok` is a verdict on the machinery, not on whether there is a value, and reading
    it as one inverts the truth. On a real run `stage` came back ok=TRUE with all
    eighteen fields empty, and ok=FALSE on the only patient who had a stage — so
    filtering on `ok` selected exactly the wrong patient and said nothing about it."""
    _a_result(tmp_path, "P1", "stage", {
        "clinical_tnm": None, "stage_at_diagnosis": None,
        "rationale": "the report describes a benign core biopsy",
    }, ok=True)
    _a_result(tmp_path, "P2", "stage", {
        "stage_at_diagnosis": {"overall_group": "I", "basis": "clinical"},
    }, ok=False)

    rows = {r["patient_id"]: r for r in _rows(recipe_tables(tmp_path)["stage"])}

    assert rows["P1"]["ok"] == "TRUE" and rows["P1"]["any_value_found"] == "FALSE"
    assert rows["P2"]["ok"] == "FALSE" and rows["P2"]["any_value_found"] == "TRUE"


def test_a_rationale_alone_is_not_a_finding(tmp_path):
    """The model always writes one. Counting it would make every row look answered."""
    _a_result(tmp_path, "P1", "stage", {"rationale": "nothing stated", "confidence": 0})

    assert _rows(recipe_tables(tmp_path)["stage"])[0]["any_value_found"] == "FALSE"


def test_rows_end_with_a_bare_newline(tmp_path):
    """csv's default \\r\\n, written through write_text with no translation, left a
    trailing \\r on the last column of every row for anyone using cut or awk."""
    _a_result(tmp_path, "P1", "date_of_birth", {"date_of_birth": "1973-01-02"})

    text = recipe_tables(tmp_path)["date_of_birth"]

    assert "\r" not in text


def test_the_source_description_is_plain_ascii(tmp_path):
    """A CSV opened by double-click in Excel on macOS is decoded as the legacy system
    encoding, and a middle dot separator came out as "¬∑" in every source cell. Chart
    text has its own non-ASCII that must survive as written — bullets, degree signs, ®
    and © are all in these reports — but the separator has no business adding to it."""
    _a_result(tmp_path, "P1", "date_of_birth", {
        "date_of_birth": "1973-01-02", "evidence_chunk_id": "P1:demographics:0:TABLE"})
    _a_passage(tmp_path, "P1", "date_of_birth", "grab", [{
        "chunk_id": "P1:demographics:0:TABLE", "text": "born 1973-01-02",
        "source_file": "demographics.csv", "record_index": 0,
        "document_date": "2015-02-18", "doc_type": "Surgical Procedure"}])

    source = _rows(recipe_tables(tmp_path)["date_of_birth"])[0]["evidence_source"]

    assert source == "demographics.csv | record 0 | 2015-02-18 | Surgical Procedure"
    assert source.isascii()


def test_a_cited_passage_survives_being_opened(tmp_path):
    """These tables carry the cited passage so a clinician can read it, and a passage is
    multi-line clinical text with degree signs and bullets in it. As CSV that is a
    quoted field spanning forty physical lines, and Excel decodes the bytes as the
    legacy system encoding unless a byte-order mark tells it otherwise. A cell holds it
    as written, on one row, with no mark needed."""
    run = tmp_path / RUN
    passage = "TUMOR 2.5 cm\n  \u2022 ER 95% positive\n  \u2022 stored at -20\u00b0C\n  \u00ae Oncotype"
    blocks = run / "patients" / "P1" / "prepared_evidence_text" / "date_of_birth" / "s1"
    blocks.mkdir(parents=True)
    (blocks / "evidence_blocks.json").write_text(json.dumps(
        [{"chunk_id": "P1:path:1:0", "text": passage, "source_file": "path.csv"}]),
        encoding="utf-8")
    _a_result(run, "P1", "date_of_birth", {
        "date_of_birth": "1973-01-02", "dob_certainty": 100,
        "dob_evidence_chunk_id": "P1:path:1:0"})

    written = write_values_tables(run)
    row = _sheet(written[0])[0]

    assert row["dob_evidence_text"] == passage, "the passage came back changed"
    assert row["ok"] is True, "a real boolean, not the string TRUE/FALSE csv forces"
    assert row["dob_certainty"] == 100, "a real number, not a string"


def test_a_passage_that_starts_with_an_equals_sign_is_not_a_formula(tmp_path):
    """A chart passage is arbitrary text and a CSV cell beginning with '=' is a live
    formula in whatever opens it. openpyxl guesses the same way unless every text cell
    is pinned to text."""
    run = tmp_path / RUN
    blocks = run / "patients" / "P1" / "prepared_evidence_text" / "date_of_birth" / "s1"
    blocks.mkdir(parents=True)
    (blocks / "evidence_blocks.json").write_text(json.dumps(
        [{"chunk_id": "P1:path:1:0", "text": "=SUM(A1:A9) per the flowsheet",
          "source_file": "path.csv"}]), encoding="utf-8")
    _a_result(run, "P1", "date_of_birth",
              {"date_of_birth": "1973-01-02", "dob_evidence_chunk_id": "P1:path:1:0"})

    written = write_values_tables(run)

    import openpyxl
    ws = openpyxl.load_workbook(written[0]).active
    header = [c.value for c in ws[1]]
    cell = ws.cell(row=2, column=header.index("dob_evidence_text") + 1)
    assert cell.data_type == "s", "the passage was written as a formula"
    assert cell.value == "=SUM(A1:A9) per the flowsheet"


def _a_sealed_schema(run_root: Path, variable: str, schema: dict, version: str = "v1"):
    """Seal a recipe's output schema into the run's code bundle, the way a real run does.

    The hash on the result envelope has to match, or the table treats the bundle as
    damaged and falls back — which is the behaviour test_a_bundle_that_lost_its_schema
    covers."""
    from jr_pipeline.runtime_enforcing_safety_and_reproducibility.reproducibility.code_bundle_component_fingerprints import (  # noqa: E501
        schema_hash,
    )

    d = run_root / "code" / "recipes" / "oncology" / variable / version
    d.mkdir(parents=True, exist_ok=True)
    path = d / f"{variable}_{version}_output_schema.json"
    path.write_text(json.dumps(schema), encoding="utf-8")
    return schema_hash(path)


_STAGE_SCHEMA = {
    "type": "object",
    "additionalProperties": True,
    "properties": {
        "stage_at_diagnosis": {
            "type": "object",
            "properties": {
                "overall_group": {"oneOf": [{"type": "string"}, {"type": "null"}]},
                "evidence_chunk_id": {"type": ["string", "null"]},
            },
        },
        # The shape that produced the phantom column: an object that may be null.
        "clinical_tnm": {
            "oneOf": [
                {"type": "object", "properties": {"cT": {"type": ["string", "null"]}}},
                {"type": "null"},
            ],
        },
        "rationale": {"type": "string"},
    },
}


def _a_result_against(run_root: Path, patient: str, variable: str, data: dict,
                      schema_hash_value: str, version: str = "v1"):
    d = run_root / "patients" / patient / "extract" / variable
    d.mkdir(parents=True, exist_ok=True)
    (d / "result.json").write_text(json.dumps({
        "payload": {"ok": True, "variable": variable, "data": data,
                    "recipe": {"name": variable, "version": version}},
        "produced_by": {"schema_hash": schema_hash_value, "run_id": run_root.name},
    }), encoding="utf-8")


def test_the_header_is_the_recipes_schema_not_what_the_model_returned(tmp_path):
    """Two patients, one of whom answered a field and one of whom did not. The header
    used to be built by walking rows and appending unseen keys, so it was a different
    header for every cohort — the same recipe over a different set of patients gave a
    different column set in a different order, and two such files would not stack."""
    run = tmp_path / RUN
    digest = _a_sealed_schema(run, "stage", _STAGE_SCHEMA)
    _a_result_against(run, "P1", "stage", {"rationale": "nothing found"}, digest)
    _a_result_against(run, "P2", "stage", {
        "stage_at_diagnosis": {"overall_group": "IIA", "evidence_chunk_id": "P2:n:1:0"},
        "clinical_tnm": {"cT": "cT2"}, "rationale": "from the path report",
    }, digest)

    rows = _rows(recipe_tables(run)["stage"])

    # Every field the recipe declares has a column, on BOTH rows, whether or not this
    # patient has a value for it. An empty cell is "no value here"; a missing column
    # used to mean the same thing and does not any more.
    for row in rows:
        assert row["stage_at_diagnosis.overall_group"] is not None
        assert row["clinical_tnm.cT"] is not None
        assert row["rationale"] is not None
    assert rows[0]["clinical_tnm.cT"] == ""
    assert rows[1]["clinical_tnm.cT"] == "cT2"


def test_the_phantom_column_is_gone(tmp_path):
    """A field declared oneOf[object, null] got TWO readings in the header: the bare
    `clinical_tnm` from the patient who returned null, and `clinical_tnm.cT` from the
    one who returned the object. The bare column was empty on every row of the
    reference run and told a reader nothing."""
    run = tmp_path / RUN
    digest = _a_sealed_schema(run, "stage", _STAGE_SCHEMA)
    _a_result_against(run, "P1", "stage", {"clinical_tnm": None}, digest)
    _a_result_against(run, "P2", "stage", {"clinical_tnm": {"cT": "cT2"}}, digest)

    header = list(_rows(recipe_tables(run)["stage"])[0])

    assert "clinical_tnm" not in header
    assert "clinical_tnm.cT" in header


def test_the_same_recipe_gives_the_same_header_over_a_different_cohort(tmp_path):
    """The concat property. Two runs of one recipe over different patients have to
    produce files that stack; a header derived from the answers did not."""
    first, second = tmp_path / "20260820_105033_b74d", tmp_path / "20260820_105034_c85e"
    for run, data in ((first, {"rationale": "a"}),
                      (second, {"clinical_tnm": {"cT": "cT2"}, "rationale": "b"})):
        digest = _a_sealed_schema(run, "stage", _STAGE_SCHEMA)
        _a_result_against(run, "P1", "stage", data, digest)

    assert (list(_rows(recipe_tables(first)["stage"])[0])
            == list(_rows(recipe_tables(second)["stage"])[0]))


def test_a_key_the_schema_never_declared_does_not_widen_the_header(tmp_path):
    """Seventeen of the twenty-one recipes allow additionalProperties, so a model may
    return a key nobody declared. Giving it a column would restore the instability the
    schema-derived header exists to remove. Dropping it would be a silent loss, which
    is the one thing this table must not do."""
    run = tmp_path / RUN
    digest = _a_sealed_schema(run, "stage", _STAGE_SCHEMA)
    _a_result_against(run, "P1", "stage", {"rationale": "a"}, digest)
    _a_result_against(run, "P2", "stage",
                      {"rationale": "b", "something_the_model_volunteered": "42"}, digest)

    rows = _rows(recipe_tables(run)["stage"])

    assert "something_the_model_volunteered" not in rows[0]
    assert rows[0]["fields_not_in_the_recipe_schema"] == ""
    assert "42" in rows[1]["fields_not_in_the_recipe_schema"]


def test_a_bundle_that_lost_its_schema_says_so_rather_than_pretending(tmp_path):
    """A run whose code bundle was re-sealed after the fact still holds a schema, and
    it may not be the one that ran — `junior verify` fails such a run for exactly that.
    Deriving a header from it and saying nothing would put the wrong promise on the
    file, so the fallback is recorded where a reader will see it."""
    run = tmp_path / RUN
    _a_sealed_schema(run, "stage", _STAGE_SCHEMA)
    _a_result_against(run, "P1", "stage", {"rationale": "a"}, "sha256:not-that-schema")
    write_values_tables(run)

    metadata = _sheet(run_output_dir(run) / "run_metadata.xlsx")[0]

    assert metadata["columns_came_from"] == "a_schema_the_run_no_longer_matches"


def test_a_list_of_citations_resolves_every_passage(tmp_path):
    """Five recipes cite with a plural evidence_chunk_ids. The whole joined list was
    looked up as one key, which never matched, so every such row claimed its citations
    had never been shown to the step."""
    run = tmp_path / RUN
    blocks = run / "patients" / "P1" / "prepared_evidence_text" / "stage" / "s1"
    blocks.mkdir(parents=True)
    (blocks / "evidence_blocks.json").write_text(json.dumps([
        {"chunk_id": "P1:path:1:0", "text": "invasive ductal carcinoma",
         "source_file": "path.csv"},
        {"chunk_id": "P1:note:9:2", "text": "grade 2", "source_file": "notes.csv"},
    ]), encoding="utf-8")
    _a_result(run, "P1", "stage", {
        "overall_group": "IIA", "evidence_chunk_ids": ["P1:path:1:0", "P1:note:9:2"]})

    row = _rows(recipe_tables(run)["stage"])[0]

    assert row["evidence_chunk_ids_text"] == "invasive ductal carcinoma ;; grade 2"
    assert "path.csv" in row["evidence_chunk_ids_source"]


def test_a_row_can_be_pooled_across_sites_and_joined_to_the_exhaust(tmp_path):
    """Without these these are site-local tables: two institutions' rows for the same
    patient id are indistinguishable, and nothing joins to the NO_PHI exhaust, which is
    what running two streams is for. The surrogate is minted in extraction_outcome's
    own domain so the join needs no second derivation."""
    import os

    from jr_pipeline.runtime_infrastructure.exhaust.secret_lifecycle import run_secret
    from jr_pipeline.runtime_infrastructure.exhaust.surrogates import (
        LinkageScope,
        make_surrogate,
    )

    run = tmp_path / "CONTAINS_PHI" / "pipeline_run_receipts" / RUN
    _a_result(run, "P1", "date_of_birth", {"date_of_birth": "1973-01-02"})
    os.environ["JR_SITE_ID"], os.environ["JR_STUDY_ID"] = "site_a", "mbc_2026"
    try:
        row = _rows(recipe_tables(run)["date_of_birth"])[0]
    finally:
        del os.environ["JR_SITE_ID"], os.environ["JR_STUDY_ID"]

    assert (row["site_id"], row["study_id"]) == ("site_a", "mbc_2026")
    assert row["patient_surrogate"] == make_surrogate(
        run_secret(RUN, tmp_path), scope=LinkageScope.RUN.value,
        record_type="extraction_outcome", output_field="patient_surrogate", raw="P1")


def test_the_one_file_download_says_the_linkage_once(tmp_path):
    """site_id, study_id, patient_surrogate and run_id are the same for every recipe of
    a run, so prefixing them per variable gave date_of_birth.site_id beside
    date_of_death.site_id — four spellings of one fact, in the shape whose whole
    complaint is sparsity."""
    from jr_pipeline.runtime_infrastructure.values_table import values_as_csv

    _a_result(tmp_path, "P1", "date_of_birth", {"date_of_birth": "1973-01-02"})
    _a_result(tmp_path, "P1", "date_of_death", {"date_of_death": None})

    header = list(_rows(values_as_csv(tmp_path, shape="wide"))[0])

    assert header[:5] == ["patient_id", "site_id", "study_id", "patient_surrogate", "run_id"]
    assert not [c for c in header[5:] if c.endswith((".site_id", ".patient_surrogate"))]


def test_a_nullable_object_answered_null_is_not_reported_as_undeclared(tmp_path):
    """A field declared oneOf[object, null] has no column of its own — removing that
    was the phantom-column fix — so answering it with null leaves a bare
    `clinical_tnm: None` in the flattened row with nowhere to go. Letting it fall into
    fields_not_in_the_recipe_schema said the recipe never declared it, when the recipe
    declares it precisely as nullable. Seen on a real run: every stage row carried
    {"clinical_tnm": null, "pathologic_tnm": null} there."""
    run = tmp_path / RUN
    digest = _a_sealed_schema(run, "stage", _STAGE_SCHEMA)
    _a_result_against(run, "P1", "stage", {"clinical_tnm": None, "rationale": "none found"}, digest)

    row = _rows(recipe_tables(run)["stage"])[0]

    assert row["fields_not_in_the_recipe_schema"] == ""
    assert row["clinical_tnm.cT"] == ""


def test_the_wrong_shape_under_a_declared_field_is_still_reported(tmp_path):
    """The exemption above is for null alone. A recipe that promised an object and got
    a string has had its contract broken, and swallowing that would hide it."""
    run = tmp_path / RUN
    digest = _a_sealed_schema(run, "stage", _STAGE_SCHEMA)
    _a_result_against(run, "P1", "stage", {"clinical_tnm": "cT2 cN0", "rationale": "x"}, digest)

    row = _rows(recipe_tables(run)["stage"])[0]

    assert "cT2 cN0" in row["fields_not_in_the_recipe_schema"]


def test_the_sheet_is_named_for_the_recipe(tmp_path):
    """The tab is what a reader sees with the workbook open, and it was being picked
    back out of the filename by splitting on "_results_". The moment the filename
    stopped carrying a run id that split stopped matching, and every tab read
    "date_of_birth_results.xlsx" — the file's own name, extension and all."""
    import openpyxl

    run = tmp_path / RUN
    _a_result(run, "P1", "date_of_birth", {"date_of_birth": "1973-01-02"})

    written = write_values_tables(run)
    titles = {p.name: openpyxl.load_workbook(p).active.title for p in written}

    assert titles == {"date_of_birth_results.xlsx": "date_of_birth",
                      "run_metadata.xlsx": "run_metadata"}


def _a_result_with_a_step_trace(run_root: Path, patient: str, variable: str, data: dict,
                                answer_came_from_step: str, content_found_by_step: dict):
    d = run_root / "patients" / patient / "extract" / variable
    d.mkdir(parents=True, exist_ok=True)
    (d / "result.json").write_text(json.dumps({
        "payload": {"ok": True, "variable": variable, "data": data,
                    "answer_came_from_step": answer_came_from_step,
                    "content_found_by_step": content_found_by_step},
        "produced_by": {},
    }), encoding="utf-8")


def test_a_row_says_when_a_later_step_answered_with_less_than_an_earlier_one(tmp_path):
    """The exact shape of a real run. treatment_lines ran nine steps; classify_lines
    found two lines and refine_lines then answered {"lines": []}, which is the answer
    that wins because the last one does. The row read n_lines=0, ok=TRUE,
    any_value_found=FALSE — indistinguishable from a patient who had no treatment.

    The pipeline is not judging whether those two lines were right. It found them, they
    are not in the answer, and that is a fact about the machinery the row must carry."""
    _a_result_with_a_step_trace(
        tmp_path, "P1", "treatment_lines",
        {"lines": [], "n_lines": 0},
        answer_came_from_step="refine_lines",
        content_found_by_step={"gather_text": 0, "perioperative_systemic": 2,
                               "classify_lines": 2, "refine_lines": 0},
    )

    row = _rows(recipe_tables(tmp_path)["treatment_lines"])[0]

    assert row["any_value_found"] == "FALSE"
    assert row["answer_came_from_step"] == "refine_lines"
    assert row["steps_that_returned_nothing"] == "gather_text; refine_lines"
    assert row["earlier_steps_that_found_more"] == (
        "perioperative_systemic (2 items); classify_lines (2 items)")


def test_a_step_after_the_answering_one_is_not_called_an_overwrite(tmp_path):
    """Only a step that ran BEFORE the answering step can have been replaced by it.
    Counting a later one would accuse the recipe of losing something it never had."""
    _a_result_with_a_step_trace(
        tmp_path, "P1", "v", {"x": 1},
        answer_came_from_step="middle",
        content_found_by_step={"first": 0, "middle": 1, "last": 5},
    )

    row = _rows(recipe_tables(tmp_path)["v"])[0]

    assert row["earlier_steps_that_found_more"] == ""
    assert row["steps_that_returned_nothing"] == "first"


def test_a_clean_recipe_says_nothing_in_the_step_trace(tmp_path):
    """The columns are always present, so the files still stack; they are empty when
    every step contributed and the answer carries what was found."""
    _a_result_with_a_step_trace(
        tmp_path, "P1", "v", {"x": 1},
        answer_came_from_step="finalize",
        content_found_by_step={"extract": 1, "finalize": 1},
    )

    row = _rows(recipe_tables(tmp_path)["v"])[0]

    assert row["answer_came_from_step"] == "finalize"
    assert row["steps_that_returned_nothing"] == ""
    assert row["earlier_steps_that_found_more"] == ""


def test_an_empty_list_is_not_a_finding(tmp_path):
    """A list flattens to a joined string, so an empty one arrives as "". Step 8's rule
    is written against raw data, where an empty list is a list and correctly not a
    value; applied to the flattened cell it counted "" as one. A recipe answering with
    an empty list read any_value_found TRUE over a row with nothing in it, inverting the
    column that exists to stop that exact misreading."""
    _a_result(tmp_path, "P1", "v", {"findings": [], "n_findings": 0})

    row = _rows(recipe_tables(tmp_path)["v"])[0]

    assert row["findings"] == ""
    assert row["any_value_found"] == "FALSE"


def test_the_run_report_says_found_for_a_recipe_not_named_after_its_answer():
    """The per-variable line after a stage asked whether a recipe found anything by
    looking for a field named after the variable. Three of the eighteen recipes are —
    date_of_birth, date_of_death, date_of_diagnosis — and the other fifteen fell
    through to "read: open it to see what it holds", a display declining to answer its
    own question on most of the book.

    It cannot go back to guessing the answer field: an earlier version took the first
    name without an underscore, picked up `rationale`, and reported found for a chart
    with nothing in it. The provenance rule has no such hole — rationale is structural."""
    import click.testing

    from apps_and_interfaces.command_line_interface import _report_extracted_variables

    summary = {"variables": {
        # named after itself, with a value
        "date_of_birth": {"ok": True, "data": {"date_of_birth": "1973-01-02"}},
        # NOT named after itself: a nested object carrying a real finding
        "stage": {"ok": True, "data": {"stage_at_diagnosis": {"overall_group": "IIA"},
                                       "rationale": "from the path report"}},
        # answers with a table of records
        "pathology_table": {"ok": True, "data": {"n_source_rows": 2,
                                                 "rows": [{"specimen": "left breast"}]}},
        # ran clean and found nothing: rationale alone is not a finding
        "menopausal_status": {"ok": True, "data": {"status_at_diagnosis": None,
                                                   "rationale": "the notes do not say"}},
        "genetics_germline": {"ok": False, "data": {}},
    }}

    runner = click.testing.CliRunner()
    with runner.isolation() as (out, _err, _):
        _report_extracted_variables(summary)
        printed = out.getvalue().decode()

    assert "open it to see what it holds" not in printed
    assert "date_of_birth: found" in printed
    assert "stage: found" in printed, "a nested answer is still an answer"
    assert "pathology_table: found" in printed, "records are the finding for a row recipe"
    assert "menopausal_status: nothing found in this chart" in printed
    assert "genetics_germline: failed" in printed


def test_a_long_recipe_name_does_not_run_into_its_own_status():
    """The progress column was a fixed 22 characters, set when every recipe name fitted.
    The book now holds names up to 39, and Python pads without clipping, so
    `breast_clinical_stage_ajcc7` rendered as `breast_clinical_stage_ajcc7done`."""
    from jr_pipeline.runtime_infrastructure.extraction_progress import report_variables

    written: list[str] = []
    on_variable = report_variables(write=written.append, animated=False)
    for name in ("date_of_birth", "breast_clinical_stage_ajcc7",
                 "local_treatment_locoregional_recurrence"):
        on_variable(name, "done", 3.0)

    for name, line in zip(("date_of_birth", "breast_clinical_stage_ajcc7",
                           "local_treatment_locoregional_recurrence"), written, strict=True):
        assert f"{name} " in line, f"the name ran into the status: {line!r}"
        assert "done" in line


def test_a_field_the_recipe_names_but_does_not_describe_keeps_its_own_cell(tmp_path):
    """stage declares pretreatment_path_stage as {"type": "object"} and nothing more,
    so it gets one column. Flattening descended into it anyway and produced dotted keys
    the header had no place for: the declared column sat empty on a real run while its
    contents went to fields_not_in_the_recipe_schema, which says the recipe never named
    the field. It named it; it just did not describe it."""
    run = tmp_path / RUN
    schema = {"type": "object", "properties": {
        "shapeless": {"oneOf": [{"type": "object", "additionalProperties": True},
                                {"type": "null"}]},
        "rationale": {"type": "string"}}}
    digest = _a_sealed_schema(run, "stage", schema)
    _a_result_against(run, "P1", "stage",
                      {"shapeless": {"evidence_chunk_id": "P1:path:1:0", "value": "pT1c"},
                       "rationale": "x"}, digest)

    row = _rows(recipe_tables(run)["stage"])[0]

    assert "pT1c" in row["shapeless"], "the declared column is empty"
    assert row["fields_not_in_the_recipe_schema"] == ""


def test_a_later_step_condensing_what_an_earlier_one_gathered_is_not_a_loss(tmp_path):
    """race_ethnicity's harmonize turns seven raw demographics fields into two canonical
    ones. Counting value-bearing leaves called that "demographics_lookup found more than
    the answer holds" on every patient — noise in the one column meant to be a strong
    signal. The case worth reading is the row that looks like a silent chart while an
    earlier step found something."""
    _a_result_with_a_step_trace(
        tmp_path, "P1", "race_ethnicity", {"race": "asian", "ethnicity": "not hispanic"},
        answer_came_from_step="harmonize",
        content_found_by_step={"demographics_lookup": 7, "harmonize": 2},
    )
    _a_result_with_a_step_trace(
        tmp_path, "P2", "race_ethnicity", {"race": None, "ethnicity": None},
        answer_came_from_step="harmonize",
        content_found_by_step={"demographics_lookup": 7, "harmonize": 0},
    )

    condensed, lost = _rows(recipe_tables(tmp_path)["race_ethnicity"])

    assert condensed["earlier_steps_that_found_more"] == "", "a reduction is not a loss"
    assert lost["earlier_steps_that_found_more"] == "demographics_lookup (7 items)"
